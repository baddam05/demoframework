import openpyxl
from openpyxl.writer.excel import save_workbook


def getnorows(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getnocolms(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readdata(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,colnum).value

def writedata(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    save_workbook(file)






